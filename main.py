import os
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel, EmailStr
from typing import Optional
from io import BytesIO
from datetime import datetime

from schemas import Reflection
from database import create_document, get_documents, db

# Optional heavy libs imported lazily inside functions to keep startup light

app = FastAPI(title="Direction Companion API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def read_root():
    return {"message": "Direction Companion API is running"}


@app.get("/test")
def test_database():
    response = {
        "backend": "✅ Running",
        "database": "❌ Not Available",
        "database_url": None,
        "database_name": None,
        "connection_status": "Not Connected",
        "collections": []
    }

    try:
        if db is not None:
            response["database"] = "✅ Available"
            response["database_url"] = "✅ Set" if os.getenv("DATABASE_URL") else "❌ Not Set"
            response["database_name"] = db.name if hasattr(db, 'name') else "✅ Connected"
            response["connection_status"] = "Connected"
            try:
                collections = db.list_collection_names()
                response["collections"] = collections[:10]
                response["database"] = "✅ Connected & Working"
            except Exception as e:
                response["database"] = f"⚠️ Connected but Error: {str(e)[:80]}"
        else:
            response["database"] = "⚠️ Available but not initialized"
    except Exception as e:
        response["database"] = f"❌ Error: {str(e)[:80]}"

    return response


# ----------------------
# Reflection Endpoints
# ----------------------
class ReflectionInput(BaseModel):
    feeling: str
    area: str
    challenge: str
    desired_outcome: str
    action_timeline: str


def build_distilled_and_guidance(data: ReflectionInput) -> tuple[str, list[str], str]:
    feeling = data.feeling.strip().lower()
    area = data.area.strip().lower()
    challenge = data.challenge.strip()
    desired = data.desired_outcome.strip()
    timeline = data.action_timeline.strip().lower()

    distilled = (
        f"You’re feeling {feeling} and seeking direction in {area}. "
        f"The core block seems to be: {challenge}. "
        f"You hope to come away with: {desired}."
    )

    guidance: list[str] = []
    # Area-based suggestions
    if "career" in area:
        guidance += [
            "List 3 roles or paths that genuinely excite you.",
            "Draft a tiny experiment for each (coffee chat, 1-day trial project, or a short course).",
        ]
    if "growth" in area or "mindset" in area:
        guidance += [
            "Write a 2-sentence reframe of your current self-talk.",
            "Choose one practice to repeat daily for 7 days (journaling, meditation, or movement).",
        ]
    if "purpose" in area:
        guidance += [
            "Identify a person or group you want to positively impact this month.",
            "Describe how your strengths could serve them in one simple action.",
        ]
    if "relationship" in area:
        guidance += [
            "Note one honest feeling you haven’t voiced yet—then plan a gentle, specific share.",
            "Ask one curious question that invites deeper understanding.",
        ]
    if "decision" in area:
        guidance += [
            "Write the 2 best options. For each, list 3 pros, 3 cons, and how Future-You feels.",
            "Do a 24-hour pause after deciding, then re-check your felt sense.",
        ]

    # Timeline-based nudge
    if any(k in timeline for k in ["now", "today", "1", "one", "2", "soon", "week"]):
        guidance.append("Pick the smallest 15-minute action and schedule it in your calendar right now.")
    else:
        guidance.append("Block a 30-minute slot this week to design your next step without distractions.")

    # Feeling-based message
    uplifting = "Progress comes from gentle steps."
    if any(k in feeling for k in ["stuck", "lost", "overwhelmed", "anxious"]):
        uplifting = "You’re not behind. Even one clear, kind step is real momentum."
    if any(k in feeling for k in ["hopeful", "excited", "curious", "ready"]):
        uplifting = "Your curiosity is the signal—follow it lightly."

    return distilled, guidance, uplifting


@app.post("/api/reflections")
def create_reflection(payload: ReflectionInput):
    distilled, guidance, message = build_distilled_and_guidance(payload)
    reflection = Reflection(
        feeling=payload.feeling,
        area=payload.area,
        challenge=payload.challenge,
        desired_outcome=payload.desired_outcome,
        action_timeline=payload.action_timeline,
        distilled=distilled,
        guidance=guidance,
        message=message,
    )

    try:
        inserted_id = create_document("reflection", reflection)
        return {"id": inserted_id, "summary": reflection.model_dump()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/reflections/{rid}")
def get_reflection(rid: str):
    try:
        docs = get_documents("reflection", {"_id": {"$eq": __import__("bson").ObjectId(rid)}})
        if not docs:
            raise HTTPException(status_code=404, detail="Reflection not found")
        doc = docs[0]
        # Convert ObjectId
        doc["id"] = str(doc.pop("_id"))
        return doc
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid id or error: {str(e)}")


class EmailPayload(BaseModel):
    email: EmailStr


@app.post("/api/reflections/{rid}/email")
def email_reflection(rid: str, payload: EmailPayload):
    # For this environment, we simulate sending and store the email on the record
    try:
        # update document
        db["reflection"].update_one(
            {"_id": __import__("bson").ObjectId(rid)},
            {"$set": {"emailed_to": payload.email, "updated_at": datetime.utcnow()}},
        )
        return {"status": "queued", "message": "Email scheduled (simulated)", "to": payload.email}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid id or error: {str(e)}")


@app.get("/api/reflections/{rid}/export")
def export_reflection(rid: str, format: str = "pdf"):
    try:
        docs = get_documents("reflection", {"_id": {"$eq": __import__("bson").ObjectId(rid)}})
        if not docs:
            raise HTTPException(status_code=404, detail="Reflection not found")
        doc = docs[0]
        # Normalize
        summary = {
            "Feeling": doc.get("feeling", ""),
            "Area": doc.get("area", ""),
            "Challenge": doc.get("challenge", ""),
            "Desired Outcome": doc.get("desired_outcome", ""),
            "Action Timeline": doc.get("action_timeline", ""),
            "Distilled": doc.get("distilled", ""),
            "Guidance": "\n- " + "\n- ".join(doc.get("guidance", []) or []),
            "Message": doc.get("message", ""),
        }

        if format.lower() == "xlsx":
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Direction Summary"
            row = 1
            for k, v in summary.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=v)
                row += 1
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = f"direction-summary-{rid}.xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
        else:
            # Default: PDF using reportlab
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import inch
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter

            textobject = c.beginText()
            textobject.setTextOrigin(inch, height - inch)
            textobject.setFont("Helvetica-Bold", 16)
            textobject.textLine("Direction Summary")
            textobject.setFont("Helvetica", 11)
            textobject.textLine("")
            for k, v in summary.items():
                textobject.textLine(f"{k}:")
                for line in str(v).split("\n"):
                    textobject.textLine(f"  {line}")
                textobject.textLine("")
            c.drawText(textobject)
            c.showPage()
            c.save()
            buffer.seek(0)
            filename = f"direction-summary-{rid}.pdf"
            return StreamingResponse(
                buffer,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
